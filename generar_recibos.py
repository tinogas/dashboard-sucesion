# -*- coding: utf-8 -*-
"""
Generador de recibos de renta en PDF
Datos: hoja Movimientos de dashboard_sucesion.xlsx, filtro REGISTRO = RENTA
Diseño: 2 recibos media carta por hoja (original + COPIA), tamaño carta
Cada recibo genera su propio PDF; el registro JSON evita duplicados.
"""

import openpyxl
from reportlab.lib.pagesizes import LETTER
from reportlab.lib.units import mm
from reportlab.lib import colors
from reportlab.pdfgen import canvas
import os
import sys
import json
import datetime

# ──────────────────────────────────────────────
# Rutas de archivos
# ──────────────────────────────────────────────
if getattr(sys, 'frozen', False):
    BASE_DIR = os.path.dirname(os.path.abspath(sys.executable))
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH  = os.path.join(BASE_DIR, 'dashboard_sucesion.xlsx')
JSON_PATH  = os.path.join(BASE_DIR, 'recibos_registro.json')
PDF_DIR    = os.path.join(BASE_DIR, 'recibos')
FIRMA_PATH = os.path.join(BASE_DIR, 'firma.png')

# ──────────────────────────────────────────────
# Configuración del locatario
# ──────────────────────────────────────────────
LOCATARIO = "Fco. Florentino Gastelum López"
CIUDAD    = "Navojoa, Sonora"
TELEFONO  = "662 182 0137"

MESES = {
    '1':  'Enero',      '2':  'Febrero',   '3':  'Marzo',
    '4':  'Abril',      '5':  'Mayo',      '6':  'Junio',
    '7':  'Julio',      '8':  'Agosto',    '9':  'Septiembre',
    '10': 'Octubre',    '11': 'Noviembre', '12': 'Diciembre',
}

# ──────────────────────────────────────────────
# Conversión de número a letras (español)
# ──────────────────────────────────────────────
UNIDADES = ['', 'UN', 'DOS', 'TRES', 'CUATRO', 'CINCO', 'SEIS', 'SIETE',
            'OCHO', 'NUEVE', 'DIEZ', 'ONCE', 'DOCE', 'TRECE', 'CATORCE',
            'QUINCE', 'DIECISÉIS', 'DIECISIETE', 'DIECIOCHO', 'DIECINUEVE']
DECENAS  = ['', 'DIEZ', 'VEINTE', 'TREINTA', 'CUARENTA', 'CINCUENTA',
            'SESENTA', 'SETENTA', 'OCHENTA', 'NOVENTA']
CENTENAS = ['', 'CIENTO', 'DOSCIENTOS', 'TRESCIENTOS', 'CUATROCIENTOS',
            'QUINIENTOS', 'SEISCIENTOS', 'SETECIENTOS', 'OCHOCIENTOS', 'NOVECIENTOS']


def _cientos(n):
    if n == 100:
        return 'CIEN'
    c = n // 100
    resto = n % 100
    partes = []
    if c:
        partes.append(CENTENAS[c])
    if resto < 20:
        if resto:
            partes.append(UNIDADES[resto])
    else:
        d, u = resto // 10, resto % 10
        partes.append(DECENAS[d] if u == 0 else f'{DECENAS[d]} Y {UNIDADES[u]}')
    return ' '.join(partes)


def numero_a_letras(monto):
    try:
        monto = float(monto)
    except (TypeError, ValueError):
        return 'CERO PESOS 00/100 M.N.'

    entero   = int(monto)
    centavos = round((monto - entero) * 100)

    if entero == 0:
        letras = 'CERO'
    elif entero < 1000:
        letras = _cientos(entero)
    elif entero < 2000:
        resto  = entero % 1000
        letras = 'MIL' + (' ' + _cientos(resto) if resto else '')
    elif entero < 1_000_000:
        miles  = entero // 1000
        resto  = entero % 1000
        letras = _cientos(miles) + ' MIL'
        if resto:
            letras += ' ' + _cientos(resto)
    else:
        millones = entero // 1_000_000
        resto    = entero % 1_000_000
        letras   = (_cientos(millones) if millones < 1000 else 'UN')
        letras  += ' MILLÓN' if millones == 1 else ' MILLONES'
        if resto:
            letras += ' ' + numero_a_letras(resto).split(' PESOS')[0]

    return f'{letras} PESOS {centavos:02d}/100 M.N.'


# ──────────────────────────────────────────────
# Dimensiones del recibo
# ──────────────────────────────────────────────
ANCHO_CARTA = LETTER[0]        # 612 pt
ALTO_CARTA  = LETTER[1]        # 792 pt
ANCHO_REC   = ANCHO_CARTA
ALTO_REC    = ALTO_CARTA / 2   # 396 pt — media carta

MAR_IZQ = 18 * mm
MAR_DER = 18 * mm
MAR_TOP = 10 * mm


# ──────────────────────────────────────────────
# Dibujo de un recibo (media carta)
# ──────────────────────────────────────────────
def dibujar_recibo(c, x0, y0, data, es_copia=False, folio_num=None):
    """
    Dibuja un recibo en el bloque (x0, y0) → (x0+ANCHO_REC, y0+ALTO_REC).
    Coordenadas ReportLab: origen abajo-izquierda.
    """
    # Borde exterior
    c.setStrokeColor(colors.HexColor('#2C3E50'))
    c.setLineWidth(1.5)
    c.rect(x0 + 4, y0 + 4, ANCHO_REC - 8, ALTO_REC - 8)

    top        = y0 + ALTO_REC
    xi         = x0 + MAR_IZQ
    xd         = x0 + ANCHO_REC - MAR_DER
    ancho_util = xd - xi

    # ── Banda de título ──────────────────────────
    y           = top - MAR_TOP
    banda_alto  = 16 * mm
    c.setFillColor(colors.HexColor('#2C3E50'))
    c.rect(xi - 4, y - banda_alto, ancho_util + 8, banda_alto, fill=1, stroke=0)
    c.setFillColor(colors.white)
    c.setFont('Helvetica-Bold', 14)
    titulo = 'RECIBO DE RENTA  —  COPIA' if es_copia else 'RECIBO DE RENTA'
    c.drawCentredString(x0 + ANCHO_REC / 2, y - banda_alto * 0.62, titulo)

    # Espacio generoso entre banda y primera línea de datos
    y -= banda_alto + 8 * mm

    # ── Fecha y número de recibo ─────────────────
    mes_str      = MESES.get(str(data.get('mes', '')).strip(), str(data.get('mes', '')))
    anio_str     = str(data.get('año', '') or '').strip()
    fecha_recibo = data.get('Fecha')
    if fecha_recibo and isinstance(fecha_recibo, datetime.datetime):
        mes_nombre = MESES.get(str(fecha_recibo.month), str(fecha_recibo.month))
        fecha_str  = f'{fecha_recibo.day} de {mes_nombre} de {fecha_recibo.year}'
    else:
        fecha_str = f'{mes_str} {anio_str}'

    folio_txt = str(folio_num) if folio_num else '-'
    c.setFillColor(colors.black)
    c.setFont('Helvetica-Bold', 8)
    c.drawString(xi, y, f'{CIUDAD}  a  {fecha_str}')
    c.drawRightString(xd, y, f'No.  {folio_txt}')
    y -= 5 * mm

    # Línea separadora
    c.setStrokeColor(colors.HexColor('#2C3E50'))
    c.setLineWidth(0.6)
    c.line(xi, y, xd, y)
    y -= 6 * mm

    # ── Helper para campos con línea base ────────
    def campo_linea(etiqueta, valor, y_pos, font_size_val=10):
        c.setFont('Helvetica-Bold', 8)
        c.setFillColor(colors.HexColor('#555555'))
        c.drawString(xi, y_pos, etiqueta)
        ancho_et = c.stringWidth(etiqueta, 'Helvetica-Bold', 8) + 4
        c.setFont('Helvetica-Bold', font_size_val)
        c.setFillColor(colors.black)
        c.drawString(xi + ancho_et, y_pos, str(valor))
        c.setStrokeColor(colors.HexColor('#BBBBBB'))
        c.setLineWidth(0.3)
        c.line(xi + ancho_et - 1, y_pos - 2, xd, y_pos - 2)
        return y_pos - 7.5 * mm

    # ── Campos de datos ──────────────────────────
    inquilino = str(data.get('Inquilino', '') or '')
    y = campo_linea('Inquilino:', inquilino, y, font_size_val=10)

    monto      = data.get('Ingresos') or 0
    importe_str = f'$ {float(monto):,.2f}'
    y = campo_linea('Importe:', importe_str, y, font_size_val=10)

    letras = numero_a_letras(monto)
    fs = 8.5
    while c.stringWidth(letras, 'Helvetica-Bold', fs) > ancho_util - 52 and fs > 6:
        fs -= 0.5
    y = campo_linea('Con letra:', letras, y, font_size_val=fs)

    concepto = str(data.get('Concepto', '') or '')
    periodo  = f'{concepto} {anio_str}' if anio_str else concepto
    y = campo_linea('Por concepto de renta del mes de:', periodo, y, font_size_val=10)

    inmueble = str(data.get('Inmueble', '') or '')
    y = campo_linea('Correspondiente al inmueble:', inmueble, y, font_size_val=9)

    via_texto = str(data.get('Via Pago', '') or '').strip()
    y = campo_linea('Vía de pago:', via_texto, y, font_size_val=9)

    y -= 3 * mm

    # ── Firma ────────────────────────────────────
    firma_cx  = x0 + ANCHO_REC / 2
    img_alto  = 20 * mm
    img_ancho = 65 * mm

    # Imagen encima de la línea
    if os.path.exists(FIRMA_PATH):
        c.drawImage(
            FIRMA_PATH,
            firma_cx - img_ancho / 2,
            y - img_alto,
            width=img_ancho,
            height=img_alto,
            preserveAspectRatio=True,
            mask='auto',
        )
        y -= img_alto

    # Línea siempre visible
    c.setStrokeColor(colors.HexColor('#2C3E50'))
    c.setLineWidth(0.8)
    c.line(firma_cx - 38 * mm, y, firma_cx + 38 * mm, y)

    # Nombre y datos del locatario debajo de la línea
    c.setFont('Helvetica-Bold', 8)
    c.setFillColor(colors.HexColor('#333333'))
    c.drawCentredString(firma_cx, y - 5 * mm, LOCATARIO)
    c.setFont('Helvetica', 7)
    c.drawCentredString(firma_cx, y - 9.5 * mm, f'{CIUDAD}  ·  Tel. {TELEFONO}')


# ──────────────────────────────────────────────
# Registro JSON de recibos generados
# ──────────────────────────────────────────────
def clave_recibo(data):
    """Clave única basada en inquilino + mes + año + inmueble."""
    return '|'.join([
        str(data.get('Inquilino', '') or '').strip(),
        str(data.get('mes',       '') or '').strip(),
        str(data.get('año',       '') or '').strip(),
        str(data.get('Inmueble',  '') or '').strip(),
    ])


def cargar_registro():
    if os.path.exists(JSON_PATH):
        with open(JSON_PATH, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {'ultimo_consecutivo': 0, 'recibos': []}


def guardar_registro(registro):
    with open(JSON_PATH, 'w', encoding='utf-8') as f:
        json.dump(registro, f, ensure_ascii=False, indent=2, default=str)


# ──────────────────────────────────────────────
# Carga del Excel
# ──────────────────────────────────────────────
def cargar_rentas(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb['Movimientos']
    raw_headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    headers = []
    for h in raw_headers:
        hs = str(h or '').strip()
        if 'REGISTRO' in hs.upper():
            headers.append('REGISTRO')
        elif hs.lower().startswith('a') and len(hs) <= 5 and 'o' in hs.lower():
            headers.append('año')
        else:
            headers.append(h)

    rentas = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(v is not None for v in row):
            continue
        d = {headers[i]: row[i] for i in range(len(headers)) if i < len(row)}
        if str(d.get('REGISTRO', '') or '').strip().upper() == 'RENTA':
            rentas.append(d)
    wb.close()
    return rentas


# ──────────────────────────────────────────────
# Rutas y nombres de archivos
# ──────────────────────────────────────────────
def _limpiar_nombre(texto):
    """Elimina caracteres inválidos para nombres de carpeta/archivo en Windows."""
    for ch in r'\/:*?"<>|':
        texto = texto.replace(ch, '')
    return texto.strip()


def carpeta_recibo(data):
    """recibos/{año}/{inmueble}/"""
    anio = _limpiar_nombre(str(data.get('año', '') or '').strip()) or 'sin_anio'
    inm  = _limpiar_nombre(str(data.get('Inmueble', '') or '').strip()) or 'sin_inmueble'
    return os.path.join(PDF_DIR, anio, inm)


def nombre_pdf(data, folio_num, copia=False):
    """2026_03_1_Recibo de pago de renta del mes de {Mes} {Inquilino}[_copia].pdf"""
    anio    = str(data.get('año', '') or '').strip()
    mes_num = str(data.get('mes', '') or '').strip().zfill(2)
    mes_str = MESES.get(str(data.get('mes', '')).strip(), str(data.get('mes', '')))
    inq     = _limpiar_nombre(str(data.get('Inquilino', '') or '').strip())
    prefijo = f'{anio}_{mes_num}_{folio_num}'
    base    = f'{prefijo}_Recibo de pago de renta del mes de {mes_str} {inq}'
    sufijo  = '_copia' if copia else ''
    return f'{base}{sufijo}.pdf'


# ──────────────────────────────────────────────
# Generación de dos PDFs individuales (original y copia) en media carta
# ──────────────────────────────────────────────
TAMANIO_MEDIA_CARTA = (ANCHO_CARTA, ALTO_REC)   # 612 x 396 pt


def generar_pdfs_recibo(data, folio_num):
    """
    Crea carpeta recibos/{año}/{inmueble}/ y genera dos PDFs media carta:
      - original: Recibo de pago de renta del mes de {Mes} {Inquilino}.pdf
      - copia:    Recibo de pago de renta del mes de {Mes} {Inquilino}_copia.pdf
    Devuelve (carpeta_relativa, nombre_original, nombre_copia).
    """
    carpeta = carpeta_recibo(data)
    os.makedirs(carpeta, exist_ok=True)

    # PDF original
    nom_orig  = nombre_pdf(data, folio_num, copia=False)
    path_orig = os.path.join(carpeta, nom_orig)
    c = canvas.Canvas(path_orig, pagesize=TAMANIO_MEDIA_CARTA)
    c.setTitle(f'Recibo de Renta No. {folio_num}')
    dibujar_recibo(c, 0, 0, data, es_copia=False, folio_num=folio_num)
    c.save()

    # PDF copia
    nom_copia  = nombre_pdf(data, folio_num, copia=True)
    path_copia = os.path.join(carpeta, nom_copia)
    c = canvas.Canvas(path_copia, pagesize=TAMANIO_MEDIA_CARTA)
    c.setTitle(f'Recibo de Renta No. {folio_num} - Copia')
    dibujar_recibo(c, 0, 0, data, es_copia=True, folio_num=folio_num)
    c.save()

    # Ruta relativa para el JSON (carpeta desde PDF_DIR)
    carpeta_rel = os.path.relpath(carpeta, PDF_DIR)
    return carpeta_rel, nom_orig, nom_copia


# ──────────────────────────────────────────────
# Punto de entrada
# ──────────────────────────────────────────────
def main(reiniciar_numeracion=False):
    os.makedirs(PDF_DIR, exist_ok=True)

    if reiniciar_numeracion and os.path.exists(JSON_PATH):
        os.remove(JSON_PATH)
        print('Registro JSON eliminado. Se reasignará consecutivo desde 1.')

    # Cargar estado previo
    registro          = cargar_registro()
    claves_generadas  = {r['clave'] for r in registro['recibos']}
    ultimo_consec     = registro['ultimo_consecutivo']

    # Leer todos los registros RENTA del Excel
    todos = cargar_rentas(XLSX_PATH)

    # Filtrar solo los que NO han sido generados aún
    nuevos = [d for d in todos if clave_recibo(d) not in claves_generadas]

    if not nuevos:
        print('No hay recibos nuevos por generar.')
        sys.exit(0)

    total = len(nuevos)
    print(f'Recibos nuevos a generar: {total}')

    for i, data in enumerate(nuevos, start=1):
        ultimo_consec += 1
        folio = ultimo_consec

        carpeta_rel, nom_orig, nom_copia = generar_pdfs_recibo(data, folio)

        # Registrar en JSON inmediatamente tras cada par de PDFs
        inq  = str(data.get('Inquilino', '') or '').strip()
        mes  = str(data.get('mes',       '') or '').strip()
        anio = str(data.get('año',       '') or '').strip()

        registro['ultimo_consecutivo'] = folio
        registro['recibos'].append({
            'consecutivo':      folio,
            'clave':            clave_recibo(data),
            'inquilino':        inq,
            'mes':              mes,
            'anio':             anio,
            'inmueble':         str(data.get('Inmueble',  '') or '').strip(),
            'importe':          float(data.get('Ingresos') or 0),
            'via_pago':         str(data.get('Via Pago',  '') or '').strip(),
            'carpeta':          carpeta_rel,
            'pdf_original':     nom_orig,
            'pdf_copia':        nom_copia,
            'fecha_generacion': datetime.datetime.now().isoformat(timespec='seconds'),
        })
        guardar_registro(registro)

        print(f'  [{i}/{total}] {nom_orig}')

    print(f'\nListo. {total} recibo(s) generado(s) en: {PDF_DIR}')


if __name__ == '__main__':
    main(reiniciar_numeracion='--reset' in sys.argv)
