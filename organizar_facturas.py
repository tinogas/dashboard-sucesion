# -*- coding: utf-8 -*-
"""
Organiza facturas CFDI (PDF + XML):
  - Crea carpetas: facturas_organizadas/{año}/{Emisor}/
  - Mueve PDF y XML a su carpeta
  - Genera reporte XLSX con datos clave de cada factura
  - Cruza con Movimientos del dashboard para agregar el Inmueble
"""

import os
import sys
import shutil
import xml.etree.ElementTree as ET
from datetime import datetime, timedelta
from collections import defaultdict

import re

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    import pdfplumber
    PDF_PLUMBER_OK = True
except ImportError:
    PDF_PLUMBER_OK = False

# ──────────────────────────────────────────────
# Rutas
# ──────────────────────────────────────────────
if getattr(sys, 'frozen', False):
    BASE_DIR = os.path.dirname(os.path.abspath(sys.executable))
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
FACTURAS_DIR   = os.path.join(BASE_DIR, 'facturas')
DEST_DIR       = os.path.join(BASE_DIR, 'facturas_organizadas')
DUPLICADOS_DIR = os.path.join(FACTURAS_DIR, 'duplicados')
XLSX_ORIGEN    = os.path.join(BASE_DIR, 'dashboard_sucesion.xlsx')
REPORTE_XLSX   = os.path.join(BASE_DIR, 'reporte_facturas.xlsx')
SIN_XML_XLSX   = os.path.join(BASE_DIR, 'facturas_sin_xml.xlsx')

# Namespaces CFDI
NS = {
    'cfdi':   'http://www.sat.gob.mx/cfd/4',
    'pago20': 'http://www.sat.gob.mx/Pagos20',
    'tfd':    'http://www.sat.gob.mx/TimbreFiscalDigital',
}

TOLERANCIA_DIAS   = 5     # ventana de búsqueda en Movimientos
TOLERANCIA_MONTO  = 0.02  # 2% de diferencia permitida en importe

# ──────────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────────
def limpiar_nombre(texto, max_len=60):
    """Elimina caracteres inválidos para nombres de carpeta en Windows."""
    for ch in r'\/:*?"<>|':
        texto = texto.replace(ch, ' ')
    texto = ' '.join(texto.split())        # colapsar espacios múltiples
    return texto[:max_len].strip()


def parse_fecha(s):
    """Convierte cadena ISO a datetime, o None."""
    if not s:
        return None
    for fmt in ('%Y-%m-%dT%H:%M:%S', '%Y-%m-%d', '%Y-%m-%dT%H:%M'):
        try:
            return datetime.strptime(str(s)[:len(fmt.replace('%Y','0000').replace('%m','00').replace('%d','00'))], fmt)
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(str(s)[:19])
    except Exception:
        return None


def montos_cercanos(a, b):
    """Devuelve True si a y b difieren menos del TOLERANCIA_MONTO."""
    if a is None or b is None:
        return False
    a, b = abs(float(a)), abs(float(b))
    if a == 0 and b == 0:
        return True
    if a == 0 or b == 0:
        return False
    return abs(a - b) / max(a, b) <= TOLERANCIA_MONTO


# ──────────────────────────────────────────────
# Lectura del XML CFDI
# ──────────────────────────────────────────────
def leer_cfdi(xml_path):
    """
    Extrae los campos relevantes de un XML CFDI 4.0.
    Devuelve un dict o None si falla.
    """
    try:
        tree = ET.parse(xml_path)
        root = tree.getroot()
    except ET.ParseError as e:
        print(f'  XML malformado: {os.path.basename(xml_path)} — {e}')
        return None

    tipo  = root.get('TipoDeComprobante', '')
    fecha = root.get('Fecha', '')[:10]
    serie = root.get('Serie', '')
    folio = root.get('Folio', '')
    total = float(root.get('Total', 0) or 0)
    subtotal = float(root.get('SubTotal', 0) or 0)
    descuento = float(root.get('Descuento', 0) or 0)
    moneda   = root.get('Moneda', 'MXN')

    emisor   = root.find('cfdi:Emisor', NS)
    receptor = root.find('cfdi:Receptor', NS)

    rfc_emisor   = emisor.get('Rfc', '')   if emisor   is not None else ''
    nom_emisor   = emisor.get('Nombre', '') if emisor   is not None else ''
    rfc_receptor = receptor.get('Rfc', '') if receptor is not None else ''
    nom_receptor = receptor.get('Nombre', '') if receptor is not None else ''

    # UUID del timbre
    tfd = root.find('.//tfd:TimbreFiscalDigital', NS)
    uuid = tfd.get('UUID', '') if tfd is not None else ''

    # Concepto principal
    concepto_elem = root.find('.//cfdi:Concepto', NS)
    descripcion   = concepto_elem.get('Descripcion', '') if concepto_elem is not None else ''

    # Todos los conceptos (desglose de productos/servicios de la factura)
    conceptos = []
    for c in root.findall('.//cfdi:Concepto', NS):
        desc = (c.get('Descripcion') or '').strip()
        if desc:
            conceptos.append({
                'descripcion': desc,
                'importe':     float(c.get('Importe', 0) or 0),
            })

    # Para tipo P (pago): el monto real está en pago20:Pago.Monto
    fecha_pago, monto_pago = '', 0.0
    if tipo == 'P':
        pago = root.find('.//pago20:Pago', NS)
        if pago is not None:
            monto_pago = float(pago.get('Monto', 0) or 0)
            fecha_pago = pago.get('FechaPago', '')[:10]
        total = monto_pago   # usar el monto real del complemento de pago

    # IVA trasladado total: solo el resumen a nivel Comprobante
    # (cfdi:Impuestos hijo directo de la raíz). Cada Concepto también trae
    # su propio Traslado anidado que ya está incluido en ese resumen — usar
    # './/cfdi:Traslado' contaría el IVA doble (una vez por Concepto y otra
    # en el resumen).
    iva = 0.0
    impuestos_cbte = root.find('cfdi:Impuestos', NS)
    if impuestos_cbte is not None:
        for tras in impuestos_cbte.findall('cfdi:Traslados/cfdi:Traslado', NS):
            if tras.get('Impuesto') == '002':
                iva += float(tras.get('Importe', 0) or 0)

    anio = fecha[:4] if fecha else 'sin_año'

    return {
        'uuid':          uuid,
        'tipo':          tipo,           # I=Ingreso, E=Egreso, P=Pago
        'fecha':         fecha,
        'fecha_pago':    fecha_pago,
        'serie':         serie,
        'folio':         folio,
        'rfc_emisor':    rfc_emisor,
        'nom_emisor':    nom_emisor,
        'rfc_receptor':  rfc_receptor,
        'nom_receptor':  nom_receptor,
        'subtotal':      subtotal,
        'iva':           round(iva, 2),
        'descuento':     descuento,
        'total':         total,
        'moneda':        moneda,
        'descripcion':   descripcion,
        'conceptos':     conceptos,
        'anio':          anio,
        'xml_path':      xml_path,
    }


# ──────────────────────────────────────────────
# Extracción de datos CFDI desde texto de PDF
# ──────────────────────────────────────────────
_RE_UUID   = re.compile(r'[Ff]olio\s*[Ff]iscal[:\s]*([0-9A-Fa-f]{8}-[0-9A-Fa-f]{4}-[0-9A-Fa-f]{4}-[0-9A-Fa-f]{4}-[0-9A-Fa-f]{12})', re.I)
_RE_FECHA_ISO  = re.compile(r'(20\d{2}-\d{2}-\d{2})T')
_RE_FECHA_MX   = re.compile(r'(\d{2})[/-](\d{2})[/-](20\d{2})')   # DD-MM-YYYY
_RE_RFC_EMISOR = re.compile(r'RFC\s*e?misor[:\s]*([A-Z&Ñ]{3,4}\d{6}[A-Z0-9]{2,3})', re.I)
_RE_RFC_PLAIN  = re.compile(r'RFC[:\s]+([A-Z&Ñ]{3,4}\d{6}[A-Z0-9]{2,3})', re.I)
_RE_NOM_EMISOR = re.compile(r'[Nn]ombre\s*[Ee]misor[:\s]*([A-ZÁÉÍÓÚÑÜ ]{5,60})')
_RE_TOTAL      = re.compile(r'[Tt]otal\s*\$?\s*([\d,]+\.?\d{0,2})')
_RE_SUBTOTAL   = re.compile(r'[Ss]ub\s*[Tt]otal\s*\$?\s*([\d,]+\.?\d{0,2})')
_RE_FECHA_PRED = re.compile(r'[Ff]echa\s+de\s+emisi[oó]n\s*\n?\s*(\d{2}-\d{2}-\d{4})', re.I)


def _parse_monto(texto):
    try:
        return float(texto.replace(',', ''))
    except (ValueError, AttributeError):
        return 0.0


def leer_pdf_cfdi(pdf_path):
    """
    Extrae datos CFDI desde el texto del PDF.
    Maneja CFDIs normales y recibos prediales municipales.
    Devuelve dict compatible con leer_cfdi() o None si falla.
    """
    if not PDF_PLUMBER_OK:
        return None
    try:
        with pdfplumber.open(pdf_path) as pdf:
            texto = '\n'.join(p.extract_text() or '' for p in pdf.pages[:2])
    except Exception as e:
        print(f'\n  PDF ilegible: {os.path.basename(pdf_path)} — {e}')
        return None

    if not texto.strip():
        return None

    # ── Detectar recibo predial municipal ──────────────────────────
    if 'RECIBO OFICIAL DE PAGO' in texto.upper() or 'CUENTA CATASTRAL' in texto.upper():
        m_fecha = _RE_FECHA_PRED.search(texto)
        if not m_fecha:
            m_fecha = _RE_FECHA_MX.search(texto)
            if m_fecha:
                dd, mm, yyyy = m_fecha.group(1), m_fecha.group(2), m_fecha.group(3)
            else:
                # Año del título "RECIBO OFICIAL DE PAGO 2026"
                m_anio = re.search(r'PAGO\s+(20\d{2})', texto)
                yyyy = m_anio.group(1) if m_anio else 'sin_año'
                mm = '01'
        else:
            dd, mm, yyyy = m_fecha.group(1).split('-')

        # Predio / nombre del archivo como descripción
        m_predio = re.search(r'[Dd]irecci[oó]n\s+de\s+predio[:\s]+(.+?)(?:\n|,\s+REG)', texto)
        descripcion = m_predio.group(1).strip() if m_predio else os.path.splitext(os.path.basename(pdf_path))[0]

        # Importe
        m_imp = re.search(r'[Ii]mporte\s*\$?\s*([\d,]+\.?\d{0,2})', texto)
        total = _parse_monto(m_imp.group(1)) if m_imp else 0.0

        fecha = f'{yyyy}-{mm}-01'
        return {
            'uuid':         '',
            'tipo':         'predial',
            'fecha':        fecha,
            'fecha_pago':   '',
            'serie':        '',
            'folio':        '',
            'rfc_emisor':   'MUNICIPIO',
            'nom_emisor':   'GOBIERNO MUNICIPAL NAVOJOA',
            'rfc_receptor': 'GALF730909CN0',
            'nom_receptor': 'FRANCISCO FLORENTINO GASTELUM LOPEZ',
            'subtotal':     total,
            'iva':          0.0,
            'total':        total,
            'moneda':       'MXN',
            'descripcion':  descripcion,
            'anio':         yyyy,
            'xml_path':     '',
            '_fuente':      'pdf',
        }

    # ── CFDI normal desde PDF ───────────────────────────────────────
    # UUID
    m = _RE_UUID.search(texto)
    uuid = m.group(1).upper() if m else ''

    # Fecha (ISO tiene prioridad)
    m = _RE_FECHA_ISO.search(texto)
    if m:
        fecha = m.group(1)
    else:
        m = _RE_FECHA_MX.search(texto)
        fecha = f'{m.group(3)}-{m.group(2)}-{m.group(1)}' if m else ''

    anio = fecha[:4] if fecha else 'sin_año'

    # RFC y nombre del emisor
    m = _RE_RFC_EMISOR.search(texto)
    if not m:
        m = _RE_RFC_PLAIN.search(texto)
    rfc_emisor = m.group(1) if m else ''

    m = _RE_NOM_EMISOR.search(texto)
    if m:
        nom_emisor = ' '.join(m.group(1).split())
    else:
        # Primeras líneas suelen ser el nombre del emisor
        lineas = [l.strip() for l in texto.split('\n') if l.strip()]
        nom_emisor = lineas[0] if lineas else 'SIN_EMISOR'

    # Montos
    m = _RE_TOTAL.search(texto)
    total = _parse_monto(m.group(1)) if m else 0.0
    m = _RE_SUBTOTAL.search(texto)
    subtotal = _parse_monto(m.group(1)) if m else 0.0

    # Descripción: primera línea de Conceptos
    m = re.search(r'Descripci[oó]n\s+([A-Za-záéíóúÁÉÍÓÚÑñ ]{5,60})', texto, re.I)
    descripcion = m.group(1).strip() if m else ''

    if not fecha and not nom_emisor and not uuid:
        return None

    return {
        'uuid':         uuid,
        'tipo':         'I',
        'fecha':        fecha,
        'fecha_pago':   '',
        'serie':        '',
        'folio':        '',
        'rfc_emisor':   rfc_emisor,
        'nom_emisor':   nom_emisor or 'SIN_EMISOR',
        'rfc_receptor': 'GALF730909CN0',
        'nom_receptor': 'FRANCISCO FLORENTINO GASTELUM LOPEZ',
        'subtotal':     subtotal,
        'iva':          round(total - subtotal, 2) if total > subtotal else 0.0,
        'total':        total,
        'moneda':       'MXN',
        'descripcion':  descripcion,
        'anio':         anio,
        'xml_path':     '',
        '_fuente':      'pdf',
    }


# ──────────────────────────────────────────────
# Cargar Movimientos del Excel
# ──────────────────────────────────────────────
def cargar_movimientos():
    """Lee la hoja Movimientos de dashboard_sucesion.xlsx."""
    if not os.path.exists(XLSX_ORIGEN):
        print(f'AVISO: no se encontró {XLSX_ORIGEN}, no se cruzarán inmuebles.')
        return []

    import openpyxl
    wb = openpyxl.load_workbook(XLSX_ORIGEN, read_only=True)
    ws = wb['Movimientos']
    headers_raw = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]

    # Normalizar encabezados
    headers = []
    for h in headers_raw:
        hs = str(h or '').strip()
        if 'REGISTRO' in hs.upper():
            headers.append('REGISTRO')
        else:
            headers.append(hs)

    movs = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(v is not None for v in row):
            continue
        d = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
        movs.append(d)
    wb.close()
    return movs


def buscar_inmueble(cfdi, movimientos):
    """
    Busca en Movimientos una fila cuya fecha y monto coincidan con el CFDI.
    Devuelve (inmueble, score) donde score: 2=fecha+monto, 1=solo monto, 0=sin match.
    """
    fecha_cfdi = parse_fecha(cfdi.get('fecha_pago') or cfdi.get('fecha'))
    monto_cfdi = cfdi.get('total', 0)

    mejor_inm, mejor_score = '', 0

    for mov in movimientos:
        fecha_mov = None
        for campo_f in ('Fecha', 'fecha', 'FECHA'):
            v = mov.get(campo_f)
            if v is not None:
                fecha_mov = parse_fecha(str(v))
                break

        # Monto: buscar en Ingresos y |Egresos|
        ing = mov.get('Ingresos') or mov.get('ingresos') or 0
        egr = mov.get('Egresos') or mov.get('egresos') or 0
        try:
            ing = abs(float(str(ing).replace(',', '') or 0))
            egr = abs(float(str(egr).replace(',', '') or 0))
        except (ValueError, TypeError):
            ing = egr = 0

        monto_coinc = montos_cercanos(monto_cfdi, ing) or montos_cercanos(monto_cfdi, egr)
        fecha_coinc = (
            fecha_cfdi is not None and fecha_mov is not None and
            abs((fecha_cfdi - fecha_mov).days) <= TOLERANCIA_DIAS
        )

        score = (2 if fecha_coinc else 0) + (1 if monto_coinc else 0)

        if score > mejor_score:
            inm = (mov.get('Inmueble') or mov.get('inmueble') or
                   mov.get('INMUEBLE') or '')
            mejor_inm   = str(inm).strip() if inm else ''
            mejor_score = score
            if score == 3:
                break   # match perfecto, no seguir buscando

    return mejor_inm, mejor_score


# ──────────────────────────────────────────────
# Organizar archivos
# ──────────────────────────────────────────────
MESES_NOMBRE = {
    '01': 'Enero',    '02': 'Febrero',  '03': 'Marzo',
    '04': 'Abril',    '05': 'Mayo',     '06': 'Junio',
    '07': 'Julio',    '08': 'Agosto',   '09': 'Septiembre',
    '10': 'Octubre',  '11': 'Noviembre','12': 'Diciembre',
}

def carpeta_destino(cfdi):
    """facturas_organizadas/{año}/{MM_Mes}/{Emisor}/"""
    fecha  = cfdi.get('fecha', '') or ''
    mes_num = fecha[5:7] if len(fecha) >= 7 else '00'
    mes_str = f'{mes_num}_{MESES_NOMBRE.get(mes_num, "sin_mes")}'
    emisor  = limpiar_nombre(cfdi.get('nom_emisor', 'SIN_EMISOR'))
    return os.path.join(DEST_DIR, cfdi.get('anio', 'sin_año'), mes_str, emisor)


def mover_evitando_sobrescritura(src, dest):
    """
    Mueve src a dest. Si dest ya existe (misma factura reprocesada), src se
    desvía a facturas/duplicados/ en vez de sobrescribir lo que ya está
    organizado. Si el nombre también choca dentro de duplicados, se agrega
    un sufijo numérico para no perder ningún archivo.
    Devuelve (ruta_final, es_duplicado).
    """
    if os.path.exists(dest):
        os.makedirs(DUPLICADOS_DIR, exist_ok=True)
        nombre = os.path.basename(src)
        base, ext = os.path.splitext(nombre)
        destino_dup = os.path.join(DUPLICADOS_DIR, nombre)
        contador = 1
        while os.path.exists(destino_dup):
            destino_dup = os.path.join(DUPLICADOS_DIR, f'{base}_dup{contador}{ext}')
            contador += 1
        shutil.move(src, destino_dup)
        return destino_dup, True

    os.makedirs(os.path.dirname(dest), exist_ok=True)
    shutil.move(src, dest)
    return dest, False


def mover_par(xml_path, carpeta, base_uuid):
    """
    Mueve el XML y su PDF correspondiente a carpeta.
    Si alguno ya existe en destino, se desvía a facturas/duplicados/ en vez
    de sobrescribirlo (ver mover_evitando_sobrescritura).
    Devuelve (xml_dest, pdf_dest, xml_duplicado, pdf_duplicado).
    """
    os.makedirs(carpeta, exist_ok=True)

    pdf_dest = None
    pdf_dup  = False

    # Buscar PDF con mismo UUID (nombre base)
    facturas_dir = os.path.dirname(xml_path)
    candidatos_pdf = []
    for nombre in os.listdir(facturas_dir):
        if not nombre.lower().endswith('.pdf'):
            continue
        # Coincidencia directa por UUID
        base = os.path.splitext(nombre)[0]
        if base.lower() == base_uuid.lower():
            candidatos_pdf.append(nombre)
        # Caso especial: "3dbe2ea8-...pdf.pdf"
        elif base_uuid.lower() in base.lower():
            candidatos_pdf.append(nombre)

    # Mover XML (a duplicados si ya existe en destino)
    xml_dest_normal = os.path.join(carpeta, os.path.basename(xml_path))
    xml_dest, xml_dup = mover_evitando_sobrescritura(xml_path, xml_dest_normal)

    # Mover PDF si encontró par (a duplicados si ya existe en destino)
    if candidatos_pdf:
        src_pdf = os.path.join(facturas_dir, candidatos_pdf[0])
        if os.path.exists(src_pdf):
            pdf_dest_normal = os.path.join(carpeta, candidatos_pdf[0])
            pdf_dest, pdf_dup = mover_evitando_sobrescritura(src_pdf, pdf_dest_normal)

    return xml_dest, pdf_dest, xml_dup, pdf_dup


# ──────────────────────────────────────────────
# Generar reporte XLSX
# ──────────────────────────────────────────────
COLS = [
    ('UUID',             40),
    ('Tipo',              6),
    ('Fecha CFDI',       12),
    ('Fecha Pago',       12),
    ('Serie-Folio',      12),
    ('RFC Emisor',       16),
    ('Emisor',           36),
    ('RFC Receptor',     16),
    ('Receptor',         36),
    ('SubTotal',         14),
    ('IVA',              12),
    ('Total / Monto',    14),
    ('Moneda',            8),
    ('Descripción',      40),
    ('Inmueble',         28),
    ('Match',            10),
    ('Carpeta',          50),
    ('PDF',              50),
]

COLOR_HDR   = '1F4E79'
COLOR_ALT   = 'EBF5FB'
COLOR_MATCH2 = 'D5F5E3'   # fecha + monto
COLOR_MATCH1 = 'FEF9E7'   # solo monto
COLOR_MATCH0 = 'FADBD8'   # sin match


def _clave_fila(uuid, carpeta_rel, pdf_nombre):
    """Clave de deduplicación: UUID si existe, si no carpeta+archivo."""
    if uuid:
        return uuid.upper()
    return f'{carpeta_rel}|{pdf_nombre}'.lower()


def generar_xlsx(filas):
    hdr_font  = Font(bold=True, color='FFFFFF', size=10)
    hdr_fill  = PatternFill('solid', fgColor=COLOR_HDR)
    hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin      = Side(style='thin', color='CCCCCC')
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Cargar libro existente o crear uno nuevo
    if os.path.exists(REPORTE_XLSX):
        wb = openpyxl.load_workbook(REPORTE_XLSX)
        ws = wb['Facturas'] if 'Facturas' in wb.sheetnames else wb.active
        # Recopilar claves ya registradas (col 1=UUID, 17=carpeta_rel, 18=pdf)
        claves_existentes = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(v is not None for v in row):
                continue
            uuid_ex  = str(row[0] or '').strip()
            carp_ex  = str(row[16] or '').strip()
            pdf_ex   = str(row[17] or '').strip()
            claves_existentes.add(_clave_fila(uuid_ex, carp_ex, pdf_ex))
        primera_fila_nueva = ws.max_row + 1
        filas_nuevas = [
            f for f in filas
            if _clave_fila(f.get('uuid',''), f.get('carpeta_rel',''), f.get('pdf_nombre',''))
               not in claves_existentes
        ]
        print(f'  Reporte existente: {primera_fila_nueva - 2} filas previas, '
              f'{len(filas_nuevas)} nuevas (de {len(filas)} procesadas).')
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Facturas'
        primera_fila_nueva = 2
        filas_nuevas = filas

        # Encabezados solo en libro nuevo
        ws.row_dimensions[1].height = 30
        for ci, (nombre, ancho) in enumerate(COLS, 1):
            c = ws.cell(1, ci, nombre)
            c.font   = hdr_font
            c.fill   = hdr_fill
            c.alignment = hdr_align
            c.border = border
            ws.column_dimensions[get_column_letter(ci)].width = ancho

    # Agregar filas nuevas
    for ri, fila in enumerate(filas_nuevas, primera_fila_nueva):
        score = fila.get('_score', 0)
        bg = COLOR_MATCH2 if score >= 3 else COLOR_MATCH1 if score >= 1 else COLOR_MATCH0
        fill = PatternFill('solid', fgColor=bg)

        valores = [
            fila.get('uuid', ''),
            fila.get('tipo', ''),
            fila.get('fecha', ''),
            fila.get('fecha_pago', ''),
            f"{fila.get('serie','')}{fila.get('folio','')}",
            fila.get('rfc_emisor', ''),
            fila.get('nom_emisor', ''),
            fila.get('rfc_receptor', ''),
            fila.get('nom_receptor', ''),
            fila.get('subtotal', 0),
            fila.get('iva', 0),
            fila.get('total', 0),
            fila.get('moneda', ''),
            fila.get('descripcion', ''),
            fila.get('inmueble', ''),
            {3: 'Fecha+Monto', 2: 'Fecha', 1: 'Monto', 0: 'Sin match'}.get(score, ''),
            fila.get('carpeta_rel', ''),
            fila.get('pdf_nombre', ''),
        ]

        for ci, val in enumerate(valores, 1):
            c = ws.cell(ri, ci, val)
            c.fill   = fill
            c.border = border
            c.alignment = Alignment(vertical='center', wrap_text=False)
            if ci in (10, 11, 12):   # columnas de dinero
                c.number_format = '"$"#,##0.00'
                c.alignment = Alignment(horizontal='right', vertical='center')

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(COLS))}1'

    wb.save(REPORTE_XLSX)
    print(f'Reporte XLSX: {REPORTE_XLSX}')


# ──────────────────────────────────────────────
# Revisión de facturas organizadas sin XML
# ──────────────────────────────────────────────
COLS_SIN_XML = [
    ('Año',         8),
    ('Mes',         16),
    ('Emisor',      36),
    ('Archivo PDF', 50),
    ('Carpeta',     60),
]


def _generar_xlsx_sin_xml(filas):
    hdr_font  = Font(bold=True, color='FFFFFF', size=10)
    hdr_fill  = PatternFill('solid', fgColor=COLOR_HDR)
    hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin      = Side(style='thin', color='CCCCCC')
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sin XML'

    ws.row_dimensions[1].height = 24
    for ci, (nombre, ancho) in enumerate(COLS_SIN_XML, 1):
        c = ws.cell(1, ci, nombre)
        c.font      = hdr_font
        c.fill      = hdr_fill
        c.alignment = hdr_align
        c.border    = border
        ws.column_dimensions[get_column_letter(ci)].width = ancho

    if not filas:
        ws.merge_cells(start_row=2, start_column=1, end_row=2,
                        end_column=len(COLS_SIN_XML))
        c = ws.cell(2, 1, 'Todas las facturas organizadas tienen su XML correspondiente.')
        c.font      = Font(bold=True, color='1E5631')
        c.fill      = PatternFill('solid', fgColor=COLOR_MATCH2)
        c.alignment = Alignment(horizontal='center', vertical='center')
    else:
        for ri, fila in enumerate(filas, 2):
            fill = PatternFill('solid', fgColor=COLOR_MATCH0)
            valores = [fila['anio'], fila['mes'], fila['emisor'],
                       fila['pdf'], fila['carpeta_rel']]
            for ci, val in enumerate(valores, 1):
                c = ws.cell(ri, ci, val)
                c.fill      = fill
                c.border    = border
                c.alignment = Alignment(vertical='center', wrap_text=False)

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(COLS_SIN_XML))}1'

    wb.save(SIN_XML_XLSX)
    print(f'Reporte XLSX: {SIN_XML_XLSX}')


def revisar_sin_xml():
    """
    Recorre facturas_organizadas/ y detecta los PDF que no tienen su XML
    correspondiente en la misma carpeta (solo comprobante en PDF, sin el
    CFDI real). Genera facturas_sin_xml.xlsx.
    """
    print('=' * 60)
    print('  Revisión de Facturas sin XML')
    print('=' * 60)

    if not os.path.isdir(DEST_DIR):
        print(f'ERROR: no existe la carpeta {DEST_DIR}')
        return []

    filas = []
    for raiz, _, archivos in os.walk(DEST_DIR):
        if DUPLICADOS_DIR in raiz:
            continue
        pdfs = sorted(f for f in archivos if f.lower().endswith('.pdf'))
        if not pdfs:
            continue
        xmls_base = [os.path.splitext(f)[0].lower()
                     for f in archivos if f.lower().endswith('.xml')]

        for pdf in pdfs:
            base_pdf = os.path.splitext(pdf)[0].lower()
            tiene_xml = any(base_pdf == xb or xb in base_pdf for xb in xmls_base)
            if tiene_xml:
                continue

            rel = os.path.relpath(raiz, BASE_DIR)
            partes = rel.split(os.sep)
            filas.append({
                'anio':        partes[1] if len(partes) > 1 else '',
                'mes':         partes[2] if len(partes) > 2 else '',
                'emisor':      partes[3] if len(partes) > 3 else '',
                'pdf':         pdf,
                'carpeta_rel': rel,
            })

    print(f'  PDFs revisados sin su XML: {len(filas)}')
    _generar_xlsx_sin_xml(filas)
    print('=' * 60)
    return filas


# ──────────────────────────────────────────────
# Programa principal
# ──────────────────────────────────────────────
def main():
    print('=' * 60)
    print('  Organizador de Facturas CFDI')
    print('=' * 60)

    if not os.path.isdir(FACTURAS_DIR):
        print(f'ERROR: no existe la carpeta {FACTURAS_DIR}')
        sys.exit(1)

    # Cargar movimientos para cruce
    print('\nCargando Movimientos...')
    movimientos = cargar_movimientos()
    print(f'  {len(movimientos)} registros cargados.')

    # ── Recopilar todos los XMLs (raíz + subdirectorios) ──────────
    xmls_paths = []
    for raiz, dirs, archivos in os.walk(FACTURAS_DIR):
        # No re-procesar carpetas ya organizadas ni los duplicados ya apartados
        if DEST_DIR in raiz or DUPLICADOS_DIR in raiz:
            continue
        for f in sorted(archivos):
            if f.lower().endswith('.xml'):
                xmls_paths.append(os.path.join(raiz, f))

    print(f'\nXMLs encontrados: {len(xmls_paths)}')

    filas_reporte = []
    errores       = []
    ya_movidos    = set()   # rutas completas de PDFs ya movidos
    duplicados    = []      # nombres de archivos desviados a facturas/duplicados/

    # ── Fase 1: procesar XMLs ──────────────────────────────────────
    total = len(xmls_paths)
    for i, xml_path in enumerate(xmls_paths, 1):
        xml_file  = os.path.basename(xml_path)
        base_uuid = os.path.splitext(xml_file)[0]

        pct   = i / total
        lleno = int(30 * pct)
        barra = '=' * lleno + '.' * (30 - lleno)
        nom_corto = xml_file[:40] + '...' if len(xml_file) > 43 else xml_file
        sys.stdout.write(f'\r[{barra}] {pct*100:5.1f}%  {nom_corto}')
        sys.stdout.flush()

        cfdi = leer_cfdi(xml_path)
        if cfdi is None:
            errores.append(xml_file)
            continue

        inmueble, score = buscar_inmueble(cfdi, movimientos)
        cfdi['inmueble'] = inmueble
        cfdi['_score']   = score

        carpeta = carpeta_destino(cfdi)

        xml_dest, pdf_dest, xml_dup, pdf_dup = mover_par(xml_path, carpeta, base_uuid)
        if pdf_dest:
            ya_movidos.add(pdf_dest)
        if xml_dup:
            duplicados.append(os.path.basename(xml_dest))
        if pdf_dup:
            duplicados.append(os.path.basename(pdf_dest))

        cfdi['carpeta_rel'] = os.path.relpath(os.path.dirname(xml_dest), BASE_DIR)
        cfdi['pdf_nombre']  = os.path.basename(pdf_dest) if pdf_dest else ''
        filas_reporte.append(cfdi)

    print()

    # ── Fase 2: PDFs sin XML (todos los subdirectorios de facturas) ─
    pdfs_sin_xml = []
    for raiz, dirs, archivos in os.walk(FACTURAS_DIR):
        if DEST_DIR in raiz or DUPLICADOS_DIR in raiz:
            continue
        for f in sorted(archivos):
            if not f.lower().endswith('.pdf'):
                continue
            ruta = os.path.join(raiz, f)
            if ruta not in ya_movidos:
                pdfs_sin_xml.append(ruta)

    if pdfs_sin_xml:
        print(f'\nProcesando {len(pdfs_sin_xml)} PDFs sin XML...')
        for i, pdf_path in enumerate(pdfs_sin_xml, 1):
            nombre = os.path.basename(pdf_path)
            sys.stdout.write(f'\r  [{i:3d}/{len(pdfs_sin_xml)}]  {nombre[:55]}')
            sys.stdout.flush()

            cfdi = leer_pdf_cfdi(pdf_path)

            if cfdi is None:
                # No se pudo extraer nada: carpeta genérica
                carpeta = os.path.join(DEST_DIR, 'sin_datos')
                dest_normal = os.path.join(carpeta, nombre)
                dest, es_dup = mover_evitando_sobrescritura(pdf_path, dest_normal)
                if es_dup:
                    duplicados.append(os.path.basename(dest))
                continue

            inmueble, score = buscar_inmueble(cfdi, movimientos)
            cfdi['inmueble'] = inmueble
            cfdi['_score']   = score

            carpeta     = carpeta_destino(cfdi)
            dest_normal = os.path.join(carpeta, nombre)
            dest, es_dup = mover_evitando_sobrescritura(pdf_path, dest_normal)
            if es_dup:
                duplicados.append(os.path.basename(dest))

            cfdi['carpeta_rel'] = os.path.relpath(os.path.dirname(dest), BASE_DIR)
            cfdi['pdf_nombre']  = os.path.basename(dest)
            filas_reporte.append(cfdi)

        print()

    # Reporte XLSX
    print(f'\nGenerando reporte ({len(filas_reporte)} facturas)...')
    generar_xlsx(filas_reporte)

    # Resumen
    match3 = sum(1 for f in filas_reporte if f['_score'] >= 3)
    match1 = sum(1 for f in filas_reporte if 1 <= f['_score'] < 3)
    match0 = sum(1 for f in filas_reporte if f['_score'] == 0)
    print(f'\nResumen de cruce con Movimientos:')
    print(f'  Fecha + Monto:  {match3}')
    print(f'  Solo Monto:     {match1}')
    print(f'  Sin coincid.:   {match0}')
    if errores:
        print(f'  XMLs con error: {len(errores)} — {errores}')

    if duplicados:
        print(f'\nDuplicados (ya existían en destino): {len(duplicados)} — movidos a {DUPLICADOS_DIR}')
        for d in duplicados:
            print(f'    - {d}')

    print(f'\nFacturas organizadas en: {DEST_DIR}')
    print('=' * 60)


if __name__ == '__main__':
    main()
