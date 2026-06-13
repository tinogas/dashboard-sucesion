import openpyxl, os, json

BASE_DIR  = r'c:\claude proyectos\Dashboard Sucesion'
XLSX_PATH = os.path.join(BASE_DIR, 'dashboard_sucesion.xlsx')
JSON_PATH = os.path.join(BASE_DIR, 'recibos_registro.json')

wb = openpyxl.load_workbook(XLSX_PATH, read_only=True)
ws = wb['Movimientos']

raw_headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

rows_data = [r for r in ws.iter_rows(min_row=2, values_only=True) if any(v is not None for v in r)]
wb.close()

print(f'Total filas con datos: {len(rows_data)}')
print()

# Buscar el registro de 5500
print('=== BUSCANDO REGISTRO 5500 ===')
for row in rows_data:
    d = dict(zip(raw_headers, row))
    if d.get('Ingresos') == 5500 or d.get('Ingresos') == 5500.0:
        print('ENCONTRADO:', d)

print()
print('=== TODAS LAS FILAS CON REGISTRO=RENTA (mostrando clave y valor) ===')
# Simular el procesamiento de headers igual que cargar_rentas
headers = []
for h in raw_headers:
    hs = str(h or '').strip()
    if 'REGISTRO' in hs.upper():
        headers.append('REGISTRO')
    elif hs.lower().startswith('a') and len(hs) <= 5 and 'o' in hs.lower():
        headers.append('año')
    else:
        headers.append(h)

rentas = []
for row in rows_data:
    d = {headers[i]: row[i] for i in range(len(headers)) if i < len(row)}
    if str(d.get('REGISTRO', '') or '').strip().upper() == 'RENTA':
        rentas.append(d)

print(f'Total RENTA encontradas: {len(rentas)}')
print()
print('=== ULTIMA FILA (sin filtro RENTA) ===')
last = rows_data[-1]
d = {headers[i]: last[i] for i in range(len(headers)) if i < len(last)}
print(d)
print(f"  REGISTRO raw: {repr(d.get('REGISTRO'))}")
