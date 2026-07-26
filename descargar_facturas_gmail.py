# -*- coding: utf-8 -*-
"""
Revisa el Gmail de tinogas@gmail.com y descarga los adjuntos PDF/XML de
correos que puedan contener facturas (CFDI) dentro de un rango de fechas.

El rango se elige desde el panel de control (dashboard_gui.py) o por línea
de comandos:

    python descargar_facturas_gmail.py --desde 2026-07-01 --hasta 2026-07-24

Si no se indica --desde se usa FECHA_INICIO_DEFAULT; si no se indica --hasta
no hay límite superior (llega hasta el correo más reciente).

A facturas/ solo entran los CFDI en los que participa el RFC propio (--rfc,
por defecto RFC_PROPIO_DEFAULT), ya sea como emisor o como receptor: las
facturas de renta emitidas a los inquilinos cuentan igual que las recibidas.
Las de terceros se apartan en revision/otros_rfc/ sin borrarse; con --rfc ""
se descarga todo sin filtrar.

Los archivos se guardan en facturas/ para que organizar_facturas.py los
procese después.

Requiere client_secret.json (ya presente en el proyecto) con el scope
gmail.readonly. En el primer uso se abre el navegador para iniciar sesión
con tinogas@gmail.com y se guarda el token en token.json.
"""

import os
import sys
import re
import json
import base64
import time
import argparse
import xml.etree.ElementTree as ET
from datetime import datetime, timezone, timedelta, date
from email.header import decode_header

from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────────
# Configuración
# ──────────────────────────────────────────────
if getattr(sys, 'frozen', False):
    BASE_DIR = os.path.dirname(os.path.abspath(sys.executable))
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

CUENTA_GMAIL     = 'tinogas@gmail.com'
CLIENT_SECRET    = os.path.join(BASE_DIR, 'client_secret.json')
TOKEN_PATH       = os.path.join(BASE_DIR, 'token.json')
FACTURAS_DIR     = os.path.join(BASE_DIR, 'facturas')
REVISION_DIR     = os.path.join(BASE_DIR, 'revision')
OTROS_RFC_DIR    = os.path.join(REVISION_DIR, 'otros_rfc')
REGISTRO_JSON    = os.path.join(BASE_DIR, 'gmail_facturas_registro.json')
REPORTE_XLSX     = os.path.join(BASE_DIR, 'gmail_facturas_reporte.xlsx')

SCOPES = ['https://www.googleapis.com/auth/gmail.readonly']

# Fecha usada cuando no se recibe un rango (primer correo con factura del
# histórico). Normalmente la GUI manda el rango elegido por el usuario.
FECHA_INICIO_DEFAULT = '2025/09/01'

# Filtro base: cualquier correo (bandeja, archivados, etiquetas) con adjunto
# PDF o XML. No se restringe por palabra clave porque muchas facturas no traen
# "factura" en el asunto; el filtrado real de qué es CFDI válido lo hace
# organizar_facturas.py al leer cada archivo.
FILTRO_ADJUNTOS = 'has:attachment (filename:pdf OR filename:xml)'

FORMATOS_FECHA = ('%Y-%m-%d', '%Y/%m/%d', '%d/%m/%Y', '%d-%m-%Y')

# RFC propio: solo se guardan en facturas/ los CFDI donde este RFC aparece
# como emisor o como receptor. La GUI puede mandar otros (varios separados por
# coma) y si llega vacío no se filtra nada.
RFC_PROPIO_DEFAULT = 'GALF730909CN0'
RE_RFC = re.compile(r'^[A-ZÑ&]{3,4}\d{6}[A-Z\d]{3}$')

EXTENSIONES_VALIDAS = ('.pdf', '.xml')

COLOR_HDR = '1F4E79'
COLOR_ALT = 'EBF5FB'


# ──────────────────────────────────────────────
# Rango de fechas
# ──────────────────────────────────────────────
def normalizar_fecha(valor, etiqueta='fecha'):
    """Convierte un texto de fecha a date. Acepta AAAA-MM-DD, AAAA/MM/DD y
    DD/MM/AAAA; también deja pasar objetos date/datetime tal cual."""
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor

    texto = str(valor).strip()
    if not texto:
        raise ValueError(f'La {etiqueta} está vacía.')
    for fmt in FORMATOS_FECHA:
        try:
            return datetime.strptime(texto, fmt).date()
        except ValueError:
            continue
    raise ValueError(f'{etiqueta[:1].upper()}{etiqueta[1:]} inválida: "{texto}" (usa AAAA-MM-DD).')


def construir_query(fecha_inicio, fecha_fin=None):
    """Arma el query de Gmail para el rango [fecha_inicio, fecha_fin] inclusive.

    Gmail trata before: como exclusivo, por eso se le suma un día al fin: así
    los correos del último día del rango sí quedan dentro.
    """
    ini = normalizar_fecha(fecha_inicio, 'fecha inicial')
    partes = [f'after:{ini:%Y/%m/%d}']

    if fecha_fin:
        fin = normalizar_fecha(fecha_fin, 'fecha final')
        if fin < ini:
            raise ValueError(
                f'El rango está invertido: la fecha final ({fin:%Y-%m-%d}) es '
                f'anterior a la inicial ({ini:%Y-%m-%d}).'
            )
        partes.append(f'before:{fin + timedelta(days=1):%Y/%m/%d}')

    partes.append(FILTRO_ADJUNTOS)
    return ' '.join(partes)


# ──────────────────────────────────────────────
# RFC propio (emisor o receptor)
# ──────────────────────────────────────────────
def normalizar_rfcs(valor):
    """'rfc1, rfc2' (o una lista) → set de RFC en mayúsculas. Vacío = sin filtro."""
    if not valor:
        return set()
    if isinstance(valor, str):
        partes = re.split(r'[,;\s]+', valor)
    else:
        partes = list(valor)

    rfcs = set()
    for p in partes:
        rfc = str(p).strip().upper()
        if not rfc:
            continue
        if not RE_RFC.match(rfc):
            raise ValueError(f'RFC inválido: "{rfc}".')
        rfcs.add(rfc)
    return rfcs


def rfcs_de_xml(contenido):
    """(RFC emisor, RFC receptor) de un CFDI. (None, None) si no se puede leer."""
    try:
        root = ET.fromstring(contenido.lstrip(b'\xef\xbb\xbf') if isinstance(contenido, bytes)
                             else contenido)
    except ET.ParseError:
        return None, None

    emisor = receptor = None
    for el in root.iter():
        tag = el.tag.rsplit('}', 1)[-1]
        if tag == 'Emisor' and emisor is None:
            emisor = (el.get('Rfc') or el.get('rfc') or '').strip().upper() or None
        elif tag == 'Receptor' and receptor is None:
            receptor = (el.get('Rfc') or el.get('rfc') or '').strip().upper() or None
        if emisor and receptor:
            break
    return emisor, receptor


def clasificar_correo(adjuntos, rfcs_permitidos):
    """Decide a qué carpeta va un correo completo (sus adjuntos no se separan).

    Un CFDI se considera propio si el RFC aparece como emisor o como receptor:
    las facturas de renta que emites a tus inquilinos son tan tuyas como las
    que te emiten a ti.

    Devuelve (carpeta, etiqueta, rfcs_emisores, rfcs_receptores).
      Factura   → PDF + XML y, si hay filtro, tu RFC participa en el CFDI.
      Otro RFC  → en ningún CFDI del correo participa tu RFC.
      Revisión  → falta PDF o XML, o no se pudo leer el RFC habiendo filtro.
    """
    emisores, receptores = set(), set()
    for a in adjuntos:
        if not a['filename'].lower().endswith('.xml'):
            continue
        emisor, receptor = rfcs_de_xml(a['contenido'])
        if emisor:
            emisores.add(emisor)
        if receptor:
            receptores.add(receptor)
    participantes = emisores | receptores

    tiene_pdf = any(a['filename'].lower().endswith('.pdf') for a in adjuntos)
    tiene_xml = any(a['filename'].lower().endswith('.xml') for a in adjuntos)
    completo  = tiene_pdf and tiene_xml

    if rfcs_permitidos and participantes and not (participantes & rfcs_permitidos):
        return OTROS_RFC_DIR, 'Otro RFC', emisores, receptores
    if rfcs_permitidos and completo and not participantes:
        # Con filtro activo, a facturas/ solo entra lo que se pudo verificar.
        return REVISION_DIR, 'Revisión', emisores, receptores
    if completo:
        return FACTURAS_DIR, 'Factura', emisores, receptores
    return REVISION_DIR, 'Revisión', emisores, receptores


# ──────────────────────────────────────────────
# Autenticación
# ──────────────────────────────────────────────
def obtener_servicio_gmail():
    creds = None
    if os.path.exists(TOKEN_PATH):
        creds = Credentials.from_authorized_user_file(TOKEN_PATH, SCOPES)

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            if not os.path.exists(CLIENT_SECRET):
                print(f'ERROR: no se encontró {CLIENT_SECRET}')
                sys.exit(1)
            print(f'Se abrirá el navegador: inicia sesión con {CUENTA_GMAIL} '
                  f'y autoriza el acceso de solo lectura a Gmail.')
            flow = InstalledAppFlow.from_client_secrets_file(CLIENT_SECRET, SCOPES)
            creds = flow.run_local_server(port=0)

        with open(TOKEN_PATH, 'w', encoding='utf-8') as f:
            f.write(creds.to_json())

    return build('gmail', 'v1', credentials=creds)


def _con_reintentos(func, *args, **kwargs):
    """Reintenta ante errores transitorios de la API (429/5xx)."""
    espera = 1
    for intento in range(5):
        try:
            return func(*args, **kwargs)
        except HttpError as e:
            if e.resp.status in (429, 500, 502, 503) and intento < 4:
                time.sleep(espera)
                espera *= 2
                continue
            raise


# ──────────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────────
def _decodificar_header(valor):
    if not valor:
        return ''
    partes = decode_header(valor)
    resultado = []
    for texto, cod in partes:
        if isinstance(texto, bytes):
            resultado.append(texto.decode(cod or 'utf-8', errors='replace'))
        else:
            resultado.append(texto)
    return ' '.join(resultado).strip()


def _limpiar_nombre_archivo(nombre):
    for ch in r'\/:*?"<>|':
        nombre = nombre.replace(ch, ' ')
    return ' '.join(nombre.split()).strip()


def _guardar_sin_sobrescribir(contenido, nombre, carpeta):
    """Guarda contenido en carpeta/nombre; si ya existe, agrega sufijo numérico."""
    os.makedirs(carpeta, exist_ok=True)
    base, ext = os.path.splitext(nombre)
    destino = os.path.join(carpeta, nombre)
    contador = 1
    while os.path.exists(destino):
        destino = os.path.join(carpeta, f'{base}_{contador}{ext}')
        contador += 1
    with open(destino, 'wb') as f:
        f.write(contenido)
    return destino


def _extraer_adjuntos(parts, encontrados):
    """Recorre recursivamente las partes MIME buscando adjuntos PDF/XML."""
    for part in parts or []:
        filename = part.get('filename') or ''
        if filename.lower().endswith(EXTENSIONES_VALIDAS):
            body = part.get('body', {})
            if body.get('attachmentId') or body.get('data'):
                encontrados.append({
                    'filename': _limpiar_nombre_archivo(filename),
                    'attachmentId': body.get('attachmentId'),
                    'data': body.get('data'),
                })
        if part.get('parts'):
            _extraer_adjuntos(part['parts'], encontrados)


def _decodificar_b64(data):
    data = data.replace('-', '+').replace('_', '/')
    falta = len(data) % 4
    if falta:
        data += '=' * (4 - falta)
    return base64.b64decode(data)


# ──────────────────────────────────────────────
# Registro incremental (evita reprocesar correos ya vistos)
# ──────────────────────────────────────────────
def cargar_registro():
    if os.path.exists(REGISTRO_JSON):
        with open(REGISTRO_JSON, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {}


def guardar_registro(registro):
    with open(REGISTRO_JSON, 'w', encoding='utf-8') as f:
        json.dump(registro, f, ensure_ascii=False, indent=2)


# ──────────────────────────────────────────────
# Reporte XLSX
# ──────────────────────────────────────────────
COLS = [
    ('Fecha correo',  16),
    ('Remitente',     40),
    ('Asunto',        50),
    ('Archivo',       45),
    ('Tipo',           6),
    ('Carpeta',       12),
    ('Enlace Gmail',  20),
    ('RFC emisor',    16),   # al final para no desalinear reportes ya generados
    ('RFC receptor',  16),
]

COL_ENLACE = 7

COLOR_FACTURA  = 'D5F5E3'   # verde: pdf + xml juntos
COLOR_REVISION = 'FEF9E7'   # amarillo: falta pdf o xml
COLOR_OTRO_RFC = 'EAEDED'   # gris: CFDI emitido a otro RFC

COLORES_CARPETA = {
    'Factura':  COLOR_FACTURA,
    'Revisión': COLOR_REVISION,
    'Otro RFC': COLOR_OTRO_RFC,
}


def generar_reporte(filas_nuevas):
    hdr_font  = Font(bold=True, color='FFFFFF', size=10)
    hdr_fill  = PatternFill('solid', fgColor=COLOR_HDR)
    hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin      = Side(style='thin', color='CCCCCC')
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)

    if os.path.exists(REPORTE_XLSX):
        wb = openpyxl.load_workbook(REPORTE_XLSX)
        ws = wb['Correos'] if 'Correos' in wb.sheetnames else wb.active
        fila_inicio = ws.max_row + 1
        # Reportes de versiones anteriores no traen la columna de RFC.
        for ci, (nombre, ancho) in enumerate(COLS, 1):
            if ws.cell(1, ci).value != nombre:
                c = ws.cell(1, ci, nombre)
                c.font, c.fill, c.alignment, c.border = hdr_font, hdr_fill, hdr_align, border
                ws.column_dimensions[get_column_letter(ci)].width = ancho
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Correos'
        ws.row_dimensions[1].height = 26
        for ci, (nombre, ancho) in enumerate(COLS, 1):
            c = ws.cell(1, ci, nombre)
            c.font, c.fill, c.alignment, c.border = hdr_font, hdr_fill, hdr_align, border
            ws.column_dimensions[get_column_letter(ci)].width = ancho
        fila_inicio = 2

    for i, fila in enumerate(filas_nuevas):
        ri = fila_inicio + i
        fill = PatternFill('solid',
                           fgColor=COLORES_CARPETA.get(fila['carpeta'], COLOR_REVISION))
        valores = [fila['fecha'], fila['remitente'], fila['asunto'],
                   fila['archivo'], fila['tipo'], fila['carpeta'], 'Abrir correo',
                   fila.get('rfc_emisor', ''), fila.get('rfc_receptor', '')]
        for ci, val in enumerate(valores, 1):
            c = ws.cell(ri, ci, val)
            c.fill, c.border = fill, border
            c.alignment = Alignment(vertical='center', wrap_text=False)
        link_cell = ws.cell(ri, COL_ENLACE)
        link_cell.hyperlink = fila['enlace']
        link_cell.font = Font(color='1155CC', underline='single')

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(COLS))}1'
    wb.save(REPORTE_XLSX)


# ──────────────────────────────────────────────
# Programa principal
# ──────────────────────────────────────────────
def main(fecha_inicio=None, fecha_fin=None, rfc_propio=RFC_PROPIO_DEFAULT):
    """Descarga los adjuntos de los correos del rango indicado.

    fecha_inicio / fecha_fin son inclusivas y aceptan AAAA-MM-DD, AAAA/MM/DD o
    DD/MM/AAAA. Sin fecha_inicio se usa FECHA_INICIO_DEFAULT; sin fecha_fin no
    hay límite superior.

    rfc_propio limita facturas/ a los CFDI donde ese RFC participa como emisor
    o receptor (uno o varios separados por coma). Con cadena vacía o None se
    descarga todo sin filtrar.
    """
    ini = normalizar_fecha(fecha_inicio or FECHA_INICIO_DEFAULT, 'fecha inicial')
    fin = normalizar_fecha(fecha_fin, 'fecha final') if fecha_fin else None
    query = construir_query(ini, fin)
    rfcs_permitidos = normalizar_rfcs(rfc_propio)

    print('=' * 60)
    print('  Descarga de Facturas desde Gmail')
    print(f'  Cuenta: {CUENTA_GMAIL}')
    print(f'  Desde:  {ini:%Y-%m-%d}')
    print(f'  Hasta:  {fin:%Y-%m-%d}' if fin else '  Hasta:  (sin límite, hasta el correo más reciente)')
    print(f'  RFC:    {", ".join(sorted(rfcs_permitidos))}' if rfcs_permitidos
          else '  RFC:    (sin filtro, se descargan todas las facturas)')
    print('=' * 60)

    servicio = obtener_servicio_gmail()
    registro = cargar_registro()

    # ── Listar mensajes que cumplen el filtro ──────────────────────
    print(f'\nBuscando correos: {query}')
    mensajes = []
    page_token = None
    while True:
        resp = _con_reintentos(
            servicio.users().messages().list(
                userId='me', q=query, maxResults=500, pageToken=page_token
            ).execute
        )
        mensajes.extend(resp.get('messages', []))
        page_token = resp.get('nextPageToken')
        if not page_token:
            break
    print(f'  Correos encontrados: {len(mensajes)}')

    nuevos = [m for m in mensajes if m['id'] not in registro]
    ya_procesados = len(mensajes) - len(nuevos)
    print(f'  Ya procesados en corridas anteriores: {ya_procesados}')
    print(f'  Nuevos por revisar: {len(nuevos)}')

    if not nuevos:
        print('\nNada nuevo que descargar.')
        return

    filas_reporte = []
    total_pdf = total_xml = correos_otro_rfc = 0

    for i, msg_ref in enumerate(nuevos, 1):
        msg_id = msg_ref['id']
        pct = i / len(nuevos)
        lleno = int(30 * pct)
        barra = '=' * lleno + '.' * (30 - lleno)
        sys.stdout.write(f'\r[{barra}] {pct*100:5.1f}%  ({i}/{len(nuevos)})')
        sys.stdout.flush()

        msg = _con_reintentos(
            servicio.users().messages().get(userId='me', id=msg_id, format='full').execute
        )

        headers = {h['name']: h['value'] for h in msg['payload'].get('headers', [])}
        remitente = _decodificar_header(headers.get('From', ''))
        asunto    = _decodificar_header(headers.get('Subject', '(sin asunto)'))

        ts_ms = int(msg.get('internalDate', 0))
        fecha = datetime.fromtimestamp(ts_ms / 1000, tz=timezone.utc).astimezone()
        fecha_str = fecha.strftime('%Y-%m-%d %H:%M')

        payload = msg['payload']
        partes = payload.get('parts') or [payload]
        adjuntos = []
        _extraer_adjuntos(partes, adjuntos)

        # El contenido se baja antes de elegir carpeta: el RFC del receptor
        # solo se conoce leyendo el XML.
        for adj in adjuntos:
            if adj['data']:
                adj['contenido'] = _decodificar_b64(adj['data'])
            else:
                att = _con_reintentos(
                    servicio.users().messages().attachments().get(
                        userId='me', messageId=msg_id, id=adj['attachmentId']
                    ).execute
                )
                adj['contenido'] = _decodificar_b64(att['data'])

        # Un correo solo cuenta como "Factura" si trae PDF y XML juntos y el
        # CFDI está a nombre de un RFC propio; lo demás va a revisión manual.
        carpeta_dest, etiqueta_carpeta, emisores, receptores = clasificar_correo(
            adjuntos, rfcs_permitidos)
        rfc_emisor_str   = ', '.join(sorted(emisores))
        rfc_receptor_str = ', '.join(sorted(receptores))
        if etiqueta_carpeta == 'Otro RFC':
            correos_otro_rfc += 1

        archivos_guardados = []
        for adj in adjuntos:
            destino = _guardar_sin_sobrescribir(adj['contenido'], adj['filename'], carpeta_dest)
            archivos_guardados.append(os.path.basename(destino))

            tipo = 'PDF' if destino.lower().endswith('.pdf') else 'XML'
            if tipo == 'PDF':
                total_pdf += 1
            else:
                total_xml += 1

            filas_reporte.append({
                'fecha': fecha_str,
                'remitente': remitente,
                'asunto': asunto,
                'archivo': os.path.basename(destino),
                'tipo': tipo,
                'carpeta': etiqueta_carpeta,
                'enlace': f'https://mail.google.com/mail/u/0/#all/{msg_id}',
                'rfc_emisor': rfc_emisor_str,
                'rfc_receptor': rfc_receptor_str,
            })

        registro[msg_id] = {
            'fecha_procesado': datetime.now().isoformat(timespec='seconds'),
            'fecha_correo': fecha_str,
            'remitente': remitente,
            'asunto': asunto,
            'carpeta': etiqueta_carpeta,
            'rfc_emisor': rfc_emisor_str,
            'rfc_receptor': rfc_receptor_str,
            'archivos': archivos_guardados,
        }

        # Guardar registro cada 25 correos por si el proceso se interrumpe
        if i % 25 == 0:
            guardar_registro(registro)

    print()
    guardar_registro(registro)

    if filas_reporte:
        print(f'\nGenerando reporte ({len(filas_reporte)} archivos nuevos)...')
        generar_reporte(filas_reporte)
        print(f'  Reporte: {REPORTE_XLSX}')
    else:
        print('\nNo se encontraron adjuntos PDF/XML en los correos nuevos.')

    correos_factura   = sum(1 for r in registro.values() if r.get('carpeta') == 'Factura' and r['archivos'])
    correos_revision  = sum(1 for r in registro.values() if r.get('carpeta') == 'Revisión' and r['archivos'])
    total_otro_rfc    = sum(1 for r in registro.values() if r.get('carpeta') == 'Otro RFC' and r['archivos'])

    print(f'\nArchivos descargados: {total_pdf} PDF, {total_xml} XML')
    print(f'  Correos con PDF+XML completos -> {FACTURAS_DIR}  ({correos_factura} correos)')
    print(f'  Correos incompletos (falta PDF o XML) -> {REVISION_DIR}  ({correos_revision} correos)')
    if rfcs_permitidos:
        print(f'  Facturas de otro RFC -> {OTROS_RFC_DIR}  '
              f'({correos_otro_rfc} correos nuevos, {total_otro_rfc} en total)')
    print('\nSiguiente paso: ejecuta organizar_facturas.py para clasificar lo que quedó en facturas/.')
    print('Revisa manualmente la carpeta revision/ para decidir qué hacer con esos correos.')
    print('=' * 60)


def _parsear_argumentos():
    p = argparse.ArgumentParser(
        description='Descarga adjuntos PDF/XML de facturas desde Gmail.')
    p.add_argument('--desde', dest='desde', default=None,
                   help='Fecha inicial inclusive (AAAA-MM-DD). '
                        f'Por defecto {FECHA_INICIO_DEFAULT}.')
    p.add_argument('--hasta', dest='hasta', default=None,
                   help='Fecha final inclusive (AAAA-MM-DD). '
                        'Por defecto sin límite superior.')
    p.add_argument('--rfc', dest='rfc', default=RFC_PROPIO_DEFAULT,
                   help='RFC propio (cuenta como emisor o receptor); varios '
                        f'separados por coma. Por defecto {RFC_PROPIO_DEFAULT}. '
                        'Usa --rfc "" para descargar sin filtrar.')
    return p.parse_args()


if __name__ == '__main__':
    args = _parsear_argumentos()
    try:
        main(fecha_inicio=args.desde, fecha_fin=args.hasta, rfc_propio=args.rfc)
    except ValueError as e:
        print(f'ERROR: {e}')
        sys.exit(1)
