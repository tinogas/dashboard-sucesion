#!/usr/bin/env python3
"""
Conciliación bancaria BBVA vs. hoja 'movimientos' — prueba exploratoria por mes.

Cruza el detalle de movimientos del estado de cuenta BBVA (PDF, es el formato más
confiable de los tres que llegan cada mes) contra lo capturado en la hoja
'movimientos' de Google Sheets para ese mismo mes, usando monto exacto + ventana
amplia de fecha (mismo mes) como criterio de emparejamiento.

Uso: python conciliacion_bbva.py --año 2025 --mes 9
Genera: conciliacion_bbva_2025_sep.xlsx (todo lo que no coincida entre banco y
registro queda marcado en rojo en ambas hojas del reporte)

No se integra al flujo de dashboard.py — es solo para validar si el cruce tiene
sentido antes de decidir si conviene volverlo una herramienta recurrente.
"""

import argparse
import os
import re
import sys

import pandas as pd
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import dashboard

AMOUNT_RE = re.compile(r"^-?\d{1,3}(,\d{3})*\.\d{2}$")
DATE_RE = re.compile(r"^\d{2}/[A-Z]{3}$")
TRASPASO_PROPIO = "TRASPASO CUENTAS PROPIAS"

MESES_ABBR = {"ENE": 1, "FEB": 2, "MAR": 3, "ABR": 4, "MAY": 5, "JUN": 6,
              "JUL": 7, "AGO": 8, "SEP": 9, "OCT": 10, "NOV": 11, "DIC": 12}


def mes_abbr(mes):
    """ene, feb, mar... a partir del número de mes (1-12)."""
    return dashboard.MESES[mes - 1].lower()


def parse_args():
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--año", "--anio", dest="año", type=int, required=True,
                    help="Año del estado de cuenta, ej. 2025")
    p.add_argument("--mes", type=int, required=True, help="Mes del estado de cuenta (1-12)")
    args = p.parse_args()
    if not 1 <= args.mes <= 12:
        p.error("--mes debe estar entre 1 y 12")
    return args.año, args.mes


# ── 1. Parseo del estado de cuenta (PDF) ───────────────────────────────────────

def parse_estado_cuenta(pdf_path, mes_objetivo):
    """Extrae los movimientos de 'Detalle de Movimientos Realizados' usando la
    posición (x0) de cada palabra para saber a qué columna (CARGOS/ABONOS/saldos)
    pertenece — el texto plano del PDF por sí solo no distingue cargo de abono."""
    movimientos = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            words = page.extract_words(use_text_flow=False, keep_blank_chars=False)
            if not words:
                continue

            lineas = {}
            for w in words:
                lineas.setdefault(round(w["top"], 1), []).append(w)
            tops = sorted(lineas.keys())

            header, header_top = None, None
            for top in tops:
                toks = sorted(lineas[top], key=lambda w: w["x0"])
                textos = [t["text"].upper() for t in toks]
                if "CARGOS" in textos and "ABONOS" in textos:
                    header, header_top = toks, top
                    break
            if header is None:
                continue  # página sin tabla de movimientos (carátula, glosario, etc.)

            cargos_x0 = next(t["x0"] for t in header if t["text"].upper() == "CARGOS")
            abonos_x0 = next(t["x0"] for t in header if t["text"].upper() == "ABONOS")
            saldos = sorted((t for t in header if t["x0"] > abonos_x0 + 5), key=lambda w: w["x0"])
            saldoop_x0 = saldos[0]["x0"] if saldos else abonos_x0 + 60
            limite_cargo_abono = (cargos_x0 + abonos_x0) / 2
            limite_abono_saldo = (abonos_x0 + saldoop_x0) / 2

            actual = None
            for top in tops:
                if top <= header_top:
                    continue
                toks = sorted(lineas[top], key=lambda w: w["x0"])
                texto_linea = " ".join(t["text"] for t in toks)
                if "Total" in texto_linea and "Movimientos" in texto_linea:
                    break  # fin de la tabla en esta página

                primero = toks[0]
                if DATE_RE.match(primero["text"]) and len(toks) > 1 and DATE_RE.match(toks[1]["text"]):
                    partes_desc = []
                    cargo = abono = None
                    for t in toks[2:]:
                        txt = t["text"]
                        if AMOUNT_RE.match(txt):
                            val = float(txt.replace(",", ""))
                            if t["x0"] < limite_cargo_abono:
                                cargo = val
                            elif t["x0"] < limite_abono_saldo:
                                abono = val
                            # columnas de saldo operación/liquidación: se ignoran
                        else:
                            partes_desc.append(txt)

                    dd, mes_abbr = primero["text"].split("/")
                    mes_num = MESES_ABBR.get(mes_abbr, mes_objetivo)
                    actual = {
                        "fecha_oper": primero["text"],
                        "fecha_liq": toks[1]["text"],
                        "dia": int(dd),
                        "mes": mes_num,
                        "descripcion": " ".join(partes_desc),
                        "cargo": cargo,
                        "abono": abono,
                        "es_traspaso_propio": TRASPASO_PROPIO in " ".join(partes_desc).upper(),
                    }
                    movimientos.append(actual)
                elif actual is not None:
                    actual["descripcion"] += " " + texto_linea

    return movimientos


def extraer_totales_declarados(pdf_path):
    """Lee los totales que el propio estado de cuenta imprime, para validar el parseo."""
    texto = ""
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            texto += (page.extract_text() or "") + "\n"

    def _num(patron):
        m = re.search(patron, texto)
        return float(m.group(1).replace(",", "")) if m else None

    def _int(patron):
        m = re.search(patron, texto)
        return int(m.group(1)) if m else None

    return {
        "cargos_monto": _num(r"TOTAL IMPORTE CARGOS\s+([\d,]+\.\d{2})"),
        "abonos_monto": _num(r"TOTAL IMPORTE ABONOS\s+([\d,]+\.\d{2})"),
        "cargos_n": _int(r"TOTAL MOVIMIENTOS CARGOS\s+(\d+)"),
        "abonos_n": _int(r"TOTAL MOVIMIENTOS ABONOS\s+(\d+)"),
    }


# ── 2. Carga de la hoja 'movimientos' (Google Sheets, con respaldo local) ─────

def cargar_movimientos_registrados(año, mes):
    try:
        df = dashboard.download_movimientos()
        print("  Descargado en vivo desde Google Sheets.")
    except Exception as exc:
        print(f"  AVISO: no se pudo descargar en vivo ({exc}).")
        print("  Usando la copia local de 'dashboard_sucesion.xlsx' (hoja Movimientos)...")
        df = pd.read_excel("dashboard_sucesion.xlsx", sheet_name="Movimientos", dtype=str).fillna("")

    df = dashboard.prepare_año_mes(df)
    fecha_col, ing_col, egr_col, conc_col, reg_col, inq_col, inm_col, _, _ = dashboard.detect_columns(df)

    df["_ingreso"] = dashboard.clean_numeric(df[ing_col]) if ing_col else 0.0
    df["_egreso"] = dashboard.clean_numeric(df[egr_col]) if egr_col else 0.0

    filtrado = df[(df["_año"] == año) & (df["_mes"] == mes)].copy()
    cols_info = {
        "conc_col": conc_col, "reg_col": reg_col,
        "inq_col": inq_col, "inm_col": inm_col, "fecha_col": fecha_col,
    }
    return filtrado, cols_info


# ── 3. Emparejamiento ──────────────────────────────────────────────────────────

def conciliar(movimientos_banco, df_registro, cols_info, año):
    """Empareja cada movimiento del banco con una fila de 'movimientos' del mismo
    mes cuyo monto coincida exactamente. Greedy: consume la fila candidata más
    cercana en fecha para no reusarla en otro match. Lo que no encuentra pareja
    de un lado ni del otro queda marcado para que se note en el Excel."""
    usados = set()
    fecha_col = cols_info["fecha_col"]
    fechas_registro = dashboard.parse_dates(df_registro[fecha_col]) if fecha_col else None

    resultados = []
    for mov in sorted(movimientos_banco, key=lambda m: m["dia"]):
        if mov["es_traspaso_propio"]:
            resultados.append({**mov, "estado": "TRASPASO PROPIO", "match": None})
            continue

        monto = mov["abono"] if mov["abono"] is not None else mov["cargo"]
        col_monto = "_ingreso" if mov["abono"] is not None else "_egreso"

        # Egreso suele guardarse en negativo en la hoja de registro; comparamos
        # por valor absoluto para no depender de esa convención de signo.
        candidatos = df_registro[
            (~df_registro.index.isin(usados)) &
            (df_registro[col_monto].abs().round(2) == round(abs(monto), 2))
        ]

        if candidatos.empty:
            resultados.append({**mov, "estado": "SIN REGISTRAR", "match": None})
            continue

        if len(candidatos) > 1 and fechas_registro is not None:
            fecha_mov = pd.Timestamp(año, mov["mes"], mov["dia"])
            diffs = (fechas_registro.loc[candidatos.index] - fecha_mov).abs()
            idx_match = diffs.sort_values().index[0]
        else:
            idx_match = candidatos.index[0]

        usados.add(idx_match)
        fila = df_registro.loc[idx_match]
        resultados.append({
            **mov,
            "estado": "CONCILIADO",
            "match": {
                "concepto": fila.get(cols_info["conc_col"], "") if cols_info["conc_col"] else "",
                "registro": fila.get(cols_info["reg_col"], "") if cols_info["reg_col"] else "",
                "inquilino": fila.get(cols_info["inq_col"], "") if cols_info["inq_col"] else "",
                "inmueble": fila.get(cols_info["inm_col"], "") if cols_info["inm_col"] else "",
            },
        })

    sin_banco = df_registro[~df_registro.index.isin(usados)]
    return resultados, sin_banco


# ── 4. Reporte Excel ───────────────────────────────────────────────────────────

FILL_OK = PatternFill("solid", fgColor="C6EFCE")
FILL_FALTA = PatternFill("solid", fgColor="FFC7CE")
FILL_TRASPASO = PatternFill("solid", fgColor="D9D9D9")
FONT_HEADER = Font(bold=True, color="FFFFFF")
FILL_HEADER = PatternFill("solid", fgColor="1F4E79")
FONT_FALTA = Font(color="9C0006", bold=True)


def _escribir_encabezado(ws, encabezados):
    for ci, texto in enumerate(encabezados, start=1):
        c = ws.cell(row=1, column=ci, value=texto)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"


def _ajustar_anchos(ws, anchos):
    for ci, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = ancho


def generar_reporte(resultados, sin_banco, cols_info, totales_declarados, año, mes, output_path):
    wb = Workbook()

    # -- Hoja Banco: cada movimiento del estado de cuenta, con su estado de conciliación --
    ws = wb.active
    ws.title = "Banco"
    encabezados = ["Fecha Oper", "Fecha Liq", "Descripción", "Cargo", "Abono",
                   "Estado", "Concepto registrado", "Registro", "Inquilino/Inmueble"]
    _escribir_encabezado(ws, encabezados)

    fill_por_estado = {"CONCILIADO": FILL_OK, "SIN REGISTRAR": FILL_FALTA, "TRASPASO PROPIO": FILL_TRASPASO}
    for ri, r in enumerate(resultados, start=2):
        match = r.get("match") or {}
        inq_inm = " / ".join(x for x in (match.get("inquilino", ""), match.get("inmueble", "")) if x)
        fila = [
            r["fecha_oper"], r["fecha_liq"], r["descripcion"],
            r["cargo"], r["abono"], r["estado"],
            match.get("concepto", ""), match.get("registro", ""), inq_inm,
        ]
        for ci, val in enumerate(fila, start=1):
            ws.cell(row=ri, column=ci, value=val)
        fill = fill_por_estado.get(r["estado"], PatternFill())
        for ci in range(1, len(fila) + 1):
            ws.cell(row=ri, column=ci).fill = fill
        if r["estado"] == "SIN REGISTRAR":
            ws.cell(row=ri, column=6).font = FONT_FALTA
    _ajustar_anchos(ws, [11, 11, 45, 12, 12, 16, 30, 20, 25])

    # -- Hoja Registro sin banco: filas de 'movimientos' del mes que NO encontraron
    #    respaldo en el banco — misma marca roja que el lado "Banco" para que las
    #    discrepancias de ambos lados se vean con el mismo criterio visual --
    ws2 = wb.create_sheet("Registro sin banco")
    cols_mostrar = [c for c in [cols_info["fecha_col"], cols_info["conc_col"], cols_info["reg_col"],
                                 cols_info["inq_col"], cols_info["inm_col"]] if c] + ["_ingreso", "_egreso"]
    _escribir_encabezado(ws2, cols_mostrar + ["Estado"])
    for ri, (_, fila) in enumerate(sin_banco[cols_mostrar].iterrows(), start=2):
        for ci, col in enumerate(cols_mostrar, start=1):
            ws2.cell(row=ri, column=ci, value=fila[col])
        c = ws2.cell(row=ri, column=len(cols_mostrar) + 1, value="SIN MOVIMIENTO BANCARIO")
        c.font = FONT_FALTA
        for ci in range(1, len(cols_mostrar) + 2):
            ws2.cell(row=ri, column=ci).fill = FILL_FALTA
    _ajustar_anchos(ws2, [16] + [30] * (len(cols_mostrar) - 3) + [13, 13, 24])

    # -- Hoja Resumen --
    ws3 = wb.create_sheet("Resumen")
    cargos_banco = [r["cargo"] for r in resultados if r["cargo"] is not None]
    abonos_banco = [r["abono"] for r in resultados if r["abono"] is not None]
    n_conciliados = sum(1 for r in resultados if r["estado"] == "CONCILIADO")
    n_sin_registrar = sum(1 for r in resultados if r["estado"] == "SIN REGISTRAR")
    n_traspaso = sum(1 for r in resultados if r["estado"] == "TRASPASO PROPIO")

    filas_resumen = [
        ("Periodo", f"{mes_abbr(mes)}/{año}"),
        ("", ""),
        ("TOTAL CARGOS (banco, parseado)", round(sum(cargos_banco), 2)),
        ("TOTAL CARGOS (declarado en estado de cuenta)", totales_declarados["cargos_monto"]),
        ("N° cargos (parseado / declarado)", f"{len(cargos_banco)} / {totales_declarados['cargos_n']}"),
        ("", ""),
        ("TOTAL ABONOS (banco, parseado)", round(sum(abonos_banco), 2)),
        ("TOTAL ABONOS (declarado en estado de cuenta)", totales_declarados["abonos_monto"]),
        ("N° abonos (parseado / declarado)", f"{len(abonos_banco)} / {totales_declarados['abonos_n']}"),
        ("", ""),
        ("Movimientos bancarios CONCILIADOS", n_conciliados),
        ("Movimientos bancarios SIN REGISTRAR", n_sin_registrar),
        ("Movimientos de traspaso entre cuentas propias (no aplica)", n_traspaso),
        ("Filas de 'movimientos' del mes SIN MOVIMIENTO BANCARIO", len(sin_banco)),
    ]
    ws3.column_dimensions["A"].width = 55
    ws3.column_dimensions["B"].width = 20
    for ri, (etiqueta, valor) in enumerate(filas_resumen, start=1):
        c_etq = ws3.cell(row=ri, column=1, value=etiqueta)
        c_val = ws3.cell(row=ri, column=2, value=valor)
        if "SIN REGISTRAR" in etiqueta or "SIN MOVIMIENTO BANCARIO" in etiqueta:
            c_etq.font = FONT_FALTA
            c_val.font = FONT_FALTA
            if isinstance(valor, int) and valor > 0:
                c_val.fill = FILL_FALTA

    wb.save(output_path)


# ── main ────────────────────────────────────────────────────────────────────

def main():
    año, mes = parse_args()
    pdf_path = f"Estados de Cuentas BBVA/{año}/{mes:02d} {dashboard.MESES[mes - 1]}.pdf"
    output_path = f"conciliacion_bbva_{año}_{mes_abbr(mes)}.xlsx"

    if not os.path.exists(pdf_path):
        sys.exit(f"No se encontró el estado de cuenta: {pdf_path}")

    print(f"Parseando estado de cuenta: {pdf_path}")
    movimientos = parse_estado_cuenta(pdf_path, mes)
    totales_declarados = extraer_totales_declarados(pdf_path)

    cargos = [m["cargo"] for m in movimientos if m["cargo"] is not None]
    abonos = [m["abono"] for m in movimientos if m["abono"] is not None]
    print(f"  {len(movimientos)} movimientos detectados "
          f"({len(cargos)} cargos, {len(abonos)} abonos).")
    print(f"  Suma cargos parseada: {sum(cargos):,.2f}  "
          f"(declarado: {totales_declarados['cargos_monto']:,.2f})")
    print(f"  Suma abonos parseada: {sum(abonos):,.2f}  "
          f"(declarado: {totales_declarados['abonos_monto']:,.2f})")

    if (round(sum(cargos), 2) != totales_declarados["cargos_monto"] or
            round(sum(abonos), 2) != totales_declarados["abonos_monto"]):
        print("  AVISO: la suma parseada no coincide con los totales declarados por el banco. "
              "Revisa el parseo antes de confiar en el reporte.")

    print("\nDescargando hoja 'movimientos'...")
    df_registro, cols_info = cargar_movimientos_registrados(año, mes)
    print(f"  {len(df_registro)} filas de 'movimientos' para {mes_abbr(mes)}/{año}.")

    print("\nConciliando...")
    resultados, sin_banco = conciliar(movimientos, df_registro, cols_info, año)

    print(f"\nGenerando reporte: {output_path}")
    generar_reporte(resultados, sin_banco, cols_info, totales_declarados, año, mes, output_path)
    print("Listo.")


if __name__ == "__main__":
    main()
